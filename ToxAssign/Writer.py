from ToxAssign import PubChemClass
from ToxAssign import Format
from ToxAssign import Match
from ToxAssign import Merge
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas


def write_chem(cmpd, sign, format, match, pub):

    wb = Workbook()
    toxic = wb.create_sheet("toxic")
    unknown = wb.create_sheet("unknown")
    safe = wb.create_sheet("safe")
    found = wb.create_sheet("found")

    for r in format.set1:
        toxic.append([cmpd,sign,"tox 1",r])

    for r in format.set2:
        toxic.append([cmpd,sign,"tox 2",r])

    for r in format.set3:
        toxic.append([cmpd,sign,"tox 3",r])

    for r in format.set4:
        toxic.append([cmpd,sign,"tox 4",r])

    for r in match.set_unfound:
        unknown.append([r])

    for r in pub.safe:
        safe.append([r])

    for r in match.found_comp:
        found.append([r])

    wb.save(cmpd+sign+".xlsx")